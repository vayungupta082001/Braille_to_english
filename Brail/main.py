# from openpyxl import workbook, load_workbook
#
# alphabet_list = []
# wb = load_workbook("Book1.xlsx")
# ws = wb.active
#
# # taking value in list
# for i in range(36):
#     # print(ws[f'A{i+1}'].value, ws[f'b{i+1}'].value)
#     curr_alphabet = ws[f'A{i + 1}'].value
#     alphabet_list.append(curr_alphabet)
# print(alphabet_list)
#
#

from openpyxl import load_workbook

alphabet_list = []
wb = load_workbook("Book1.xlsx")
ws = wb.active

# taking value in list
for i in range(36):
    curr_alphabet = ws[f'A{i + 1}'].value
    alphabet_list.append(curr_alphabet)
print(alphabet_list)

# CODE FOR LETTERS ONLY
def eng_letters_to_braille():
    i=0
    alpha = input("Enter the Alphabet : ")
    alpha = alpha.upper()
    if alpha.upper() in alphabet_list:
        index = alphabet_list.index(alpha)
        print(ws[f'b{index + 1}'].value)


# CODE FOR WORDS ONLY
def eng_word_to_braille():

    word = input("Enter the Word : ")
    word = word.upper()
    # print(word)
    alpha = ""
    word = word.strip(" ")
    l = []
    for k in range(len(word)):
        l.append(word[k])
    # print(l)
    braille_word = ""
    for j in range(len(l)):
        if l[j] in alphabet_list:
            index = alphabet_list.index(l[j])
            curr_alphabet = ws[f'b{index + 1}'].value
            braille_word += curr_alphabet
    print(braille_word)


# ONLY FOR SENTENCES
def eng_sentence_to_braille():
    i=0
    sentence = input("Enter a sentence: ")
    sentence = sentence.upper()
    word = sentence.split()

    braille_sentence = ""

    for i in range(len(word)):
        braille_word = ""
        for j in range(len(word[i])):
            curr_word = word[i]
            if curr_word[j] in alphabet_list:
                index = alphabet_list.index(curr_word[j])
                braille_word += ws[f"b{index + 1}"].value
        braille_sentence += braille_word + " "  # Adding a space after each word
    print(braille_sentence)

while True:
    if i>2:
        print("--------------------------------------------------------------------------------------------------------------")
    operation = int(input(
        "Enter \n1)English Letters to braille\n2)English Words to braille\n3)English Sentences to braille\n4)QUIT\n"))
    i += 1
    match operation:
        case 1:
            eng_letters_to_braille()
        case 2:
            eng_word_to_braille()
        case 3:
            eng_sentence_to_braille()
        case 4:
            break